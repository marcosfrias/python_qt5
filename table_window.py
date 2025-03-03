from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import QAction, QFileDialog, QMessageBox
import openpyxl

class TableWindow(QtWidgets.QMainWindow):
    def __init__(self, tableWidget):
        super().__init__()
        uic.loadUi('table_template.ui', self)

        self.copy_and_transpose_table(tableWidget)

        # Mantener una copia del contenido original de la tabla
        self.original_content = self.copy_table_content()

        # Conectar las acciones del archivo .ui con las funciones correspondientes
        self.actionGuardar.triggered.connect(self.guardar_tabla)
        self.actionNuevo.triggered.connect(self.nuevo_documento)
        self.actionSalir.triggered.connect(self.salir_aplicacion)

    def guardar_tabla(self):
        """Guardar los datos de la tabla en un archivo Excel (.xlsx)."""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar tabla como Excel", "", "Archivos Excel (*.xlsx);;Todos los archivos (*)", options=options
        )
        if file_path:
            try:
                # Crear un nuevo libro de Excel
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Datos de la tabla"

                # Obtener datos de la tabla
                row_count = self.tableWidget.rowCount()
                column_count = self.tableWidget.columnCount()

                # Escribir encabezados en el archivo Excel
                for col in range(column_count):
                    header_item = self.tableWidget.horizontalHeaderItem(col)
                    sheet.cell(row=1, column=col + 1, value=header_item.text() if header_item else "")

                # Escribir datos de la tabla
                for row in range(row_count):
                    for col in range(column_count):
                        item = self.tableWidget.item(row, col)
                        sheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

                # Guardar el archivo Excel
                workbook.save(file_path)
                QMessageBox.information(self, "Éxito", "Tabla guardada correctamente como archivo Excel.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo guardar la tabla: {e}")

    def nuevo_documento(self):
        """Reabrir la tabla con el contenido original."""
        try:
            self.restore_table_content(self.original_content)
            QMessageBox.information(self, "Nuevo Documento", "La tabla se ha restaurado con el contenido original.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo restaurar la tabla: {e}")


    def salir_aplicacion(self):
        """Cerrar la aplicación."""
        self.close()

    def copy_and_transpose_table(self, source_table):
        rowCount = source_table.rowCount()
        columnCount = source_table.columnCount()

        # Crear listas para almacenar los nuevos encabezados y datos
        new_headers = []
        new_data = []

        # Filtrar y transponer los datos
        for j in range(columnCount):
            column_data = []
            for i in range(rowCount):
                item = source_table.item(i, j)
                if i == 0 and (item is None or item.text().strip().lower() == "nan"):
                    break
                column_data.append(item.text() if item else "")
            else:
                # Si no se encontró "NaN" en la fila 0, agregar los datos y el encabezado
                new_headers.append(
                    source_table.horizontalHeaderItem(j).text() if source_table.horizontalHeaderItem(j) else '')
                new_data.append(column_data)


        # Configurar dimensiones del nuevo TableWidget
        self.tableWidget.setRowCount(len(new_headers))
        self.tableWidget.setColumnCount(rowCount)
        self.tableWidget.setHorizontalHeaderLabels(
            [source_table.verticalHeaderItem(i).text() if source_table.verticalHeaderItem(i) else '' for i in
             range(rowCount)])
        self.tableWidget.setVerticalHeaderLabels(new_headers)

        # Copiar los datos filtrados y transpuestos al nuevo TableWidget
        for i in range(len(new_headers)):
            for j in range(rowCount):
                self.tableWidget.setItem(i, j, QtWidgets.QTableWidgetItem(new_data[i][j]))

        # Ajustar el ancho de las columnas
        for col in range(len(new_headers)):
            self.tableWidget.resizeColumnToContents(col)

    def copy_table_content(self):
        """Guardar el contenido actual de la tabla en una lista de listas."""
        content = []
        row_count = self.tableWidget.rowCount()
        column_count = self.tableWidget.columnCount()

        # Copiar encabezados horizontales
        headers = [
            self.tableWidget.horizontalHeaderItem(col).text() if self.tableWidget.horizontalHeaderItem(col) else ""
            for col in range(column_count)]
        content.append(headers)

        # Copiar el contenido de cada celda
        for row in range(row_count):
            row_data = [
                self.tableWidget.item(row, col).text() if self.tableWidget.item(row, col) else ""
                for col in range(column_count)
            ]
            content.append(row_data)

        return content

    def restore_table_content(self, content):
        """Restaurar el contenido de la tabla desde una lista de listas."""
        if not content:
            return

        # Configurar las dimensiones de la tabla
        headers = content[0]
        data = content[1:]

        self.tableWidget.setRowCount(len(data))
        self.tableWidget.setColumnCount(len(headers))
        self.tableWidget.setHorizontalHeaderLabels(headers)

        # Rellenar las celdas con los datos
        for row, row_data in enumerate(data):
            for col, value in enumerate(row_data):
                self.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(value))